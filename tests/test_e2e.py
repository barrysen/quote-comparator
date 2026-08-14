"""端到端回归（设计方案 §7）：

- 3 份 Excel/文本虚拟样本经完整流水线（replay 驱动阶段三）提取正确；
- 准确率口径：字段准确率 = 1 - (值错误且未标黄的关键字段数 / 关键字段总数)，要求 ≥ 95%；
- 空文件夹跑通生成空报告（M1 完成标志）。
"""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.cli import main

PROJECT_ROOT = Path(__file__).parent.parent
FIXTURES = PROJECT_ROOT / "tests" / "fixtures"

KEY_FIELDS = ("quantity", "unit_price")  # item_name 通过匹配本身校验

OCR_FILES = {"鑫源自动化.pdf", "广联轴承.png"}
DIGITAL_FILES = {"华东机电.xlsx", "铭泰五交化.txt", "华东机电-补充.txt", "恒力工贸.pdf"}
ALL_FILES = DIGITAL_FILES | OCR_FILES


def _ocr_available() -> bool:
    import importlib.util
    find = importlib.util.find_spec
    return bool(
        (find("paddleocr") and find("paddle"))
        or (find("rapidocr_onnxruntime") and find("onnxruntime"))
    )


def _load_result(out: Path):
    data = json.loads(out.with_suffix(".result.json").read_text(encoding="utf-8"))
    return {q["source_file"]: q for q in data["quotes"]}


def _check_file(quotes: dict, expected: dict) -> tuple[int, int]:
    """返回 (关键字段总数, 错且未标黄数)。设计口径：错了但标黄不算事故。"""
    total, bad = 0, 0
    quote = quotes[expected["source_file"]]
    for exp_line in expected["lines"]:
        total += 1  # item_name 计 1 个关键字段
        matches = [
            l for l in quote["lines"]
            if l["item_name"] == exp_line["item_name"]
        ]
        if not matches:
            bad += 1
            continue
        line = matches[0]
        if exp_line.get("spec_contains"):
            assert exp_line["spec_contains"] in line["spec"], (
                f"{line['item_name']} 规格 {line['spec']!r} 应包含 {exp_line['spec_contains']!r}"
            )
        yellow = line["confidence"] == "low" or bool(line["flags"])
        for field in KEY_FIELDS:
            exp_val = exp_line.get(field)
            if exp_val is None:
                continue
            total += 1
            got = line.get(field)
            ok = got is not None and abs(got - exp_val) <= max(abs(exp_val) * 1e-6, 1e-9)
            if not ok and not yellow:
                bad += 1
    return total, bad


class TestE2E:
    @pytest.fixture()
    def result(self, tmp_path, monkeypatch):
        monkeypatch.chdir(PROJECT_ROOT)  # settings 内相对路径基于项目根
        out = tmp_path / "比价结果.xlsx"
        code = main([
            "extract", str(FIXTURES / "files"),
            "-o", str(out),
            "--config", str(FIXTURES / "settings_test.yml"),
            "--force",
        ])
        assert code == 0
        return out

    def test_outputs_exist(self, result):
        assert result.exists()
        assert result.with_suffix(".result.json").exists()
        assert result.with_suffix(".report.txt").exists()
        wb = load_workbook(result)
        assert {"明细数据", "统计与异常"} <= set(wb.sheetnames)

    def test_all_files_parsed(self, result):
        quotes = _load_result(result)
        assert set(quotes) == ALL_FILES
        for name, q in quotes.items():
            assert not [w for w in q["warnings"] if w.startswith("LLM_FAILED")], name
            if name in OCR_FILES and not _ocr_available():
                # OCR 引擎未安装：优雅降级（空结果 + 警告），不算失败
                assert q["lines"] == []
                assert q["warnings"]
            else:
                assert q["lines"], f"{name} 应有报价行"

    def test_supplier_names(self, result):
        quotes = _load_result(result)
        for exp_file in sorted((FIXTURES / "expected").glob("*.json")):
            exp = json.loads(exp_file.read_text(encoding="utf-8"))
            name = exp["source_file"]
            if name in OCR_FILES and not _ocr_available():
                continue  # 降级文件供应商名回退为文件名，不参与本断言
            assert exp["supplier_name_contains"] in quotes[name]["supplier_name"]

    def test_key_field_accuracy(self, result):
        """关键字段（物料名、单价、数量）提取准确率 ≥ 95%，错且未标黄才算事故。"""
        quotes = _load_result(result)
        total, bad = 0, 0
        for exp_file in sorted((FIXTURES / "expected").glob("*.json")):
            exp = json.loads(exp_file.read_text(encoding="utf-8"))
            if not quotes[exp["source_file"]]["lines"]:
                continue  # 降级文件无报价行，不计入准确率
            t, b = _check_file(quotes, exp)
            total += t
            bad += b
        accuracy = 1 - bad / total if total else 1.0
        print(f"\n关键字段准确率: {accuracy:.2%}（{total - bad}/{total}，错且未标黄 {bad}）")
        assert accuracy >= 0.95

    def test_merged_supplier_files(self, result):
        """同一供应商两个文件都能解析，且补充报价在比价中覆盖主报价。"""
        quotes = _load_result(result)
        assert "华东机电" in quotes["华东机电-补充.txt"]["supplier_name"]
        assert len(quotes["华东机电-补充.txt"]["lines"]) == 2
        data = json.loads(result.with_suffix(".result.json").read_text(encoding="utf-8"))
        assert any("覆盖" in n for n in data["comparison"]["notes"])

    def test_comparison_groups(self, result):
        """比价矩阵：8 种物料全部跨供应商对齐；恒力缺报气缸被捕获。"""
        wb = load_workbook(result)
        assert "比价总表" in wb.sheetnames
        assert wb.sheetnames[0] == "比价总表"
        data = json.loads(result.with_suffix(".result.json").read_text(encoding="utf-8"))
        cmp = data["comparison"]
        expected_suppliers = {"华东机电设备有限公司", "铭泰五交化", "恒力工贸有限公司"}
        if _ocr_available():
            expected_suppliers |= {"鑫源自动化科技有限公司", "广联轴承销售部"}
        assert set(cmp["suppliers"]) == expected_suppliers
        assert len(cmp["groups"]) == 8
        for g in cmp["groups"]:
            assert len(g["entries"]) >= 2, f"{g['item_name']} 应有 ≥2 家报价"
            assert not g["review"]
        # 恒力工贸缺报气缸（≥2 家已报）→ 必须被缺报检测捕获
        assert any(
            "气缸" in m["item"] and "恒力" in m["supplier"] for m in cmp["missing"]
        )
        # 名称写法不同的物料被 LLM 兜底合并（如 6204轴承 ≈ 深沟球轴承 6204）
        assert any("合并" in n for n in cmp["notes"])
        if _ocr_available():
            # OCR 可用时唯一缺报就是恒力气缸
            assert len(cmp["missing"]) == 1

    def test_v3_amount_mismatch_caught(self, result):
        """华东机电样本埋的 V3 雷（轴承 50×8.5=425 写成 500）必须被捕获标黄。"""
        data = json.loads(result.with_suffix(".result.json").read_text(encoding="utf-8"))
        v3 = [f for f in data["findings"] if f["rule"] == "V3"]
        assert any("华东机电.xlsx" in f["target"] and "差额" in f["message"] for f in v3)
        quotes = _load_result(result)
        bearing = [l for l in quotes["华东机电.xlsx"]["lines"] if "轴承" in l["item_name"]][0]
        assert "V3" in bearing["flags"]

    def test_v5_lead_time_parsed(self, result):
        """模糊交期解析：'两周'→14 天，'现货'→0 天。"""
        quotes = _load_result(result)
        mingtai = quotes["铭泰五交化.txt"]["lines"]
        valve = [l for l in mingtai if "球阀" in l["item_name"]][0]
        assert valve["lead_time_days"] == 14  # 两周
        huadong = quotes["华东机电.xlsx"]["lines"]
        screw = [l for l in huadong if l["item_name"] == "内六角螺钉"][0]
        assert screw["lead_time_days"] == 0  # 现货


class TestEmptyFolder:
    def test_m1_empty_run(self, tmp_path, monkeypatch):
        monkeypatch.chdir(PROJECT_ROOT)
        folder = tmp_path / "empty"
        folder.mkdir()
        out = tmp_path / "空报告.xlsx"
        assert main([
            "extract", str(folder), "-o", str(out),
            "--config", str(FIXTURES / "settings_test.yml"), "--force",
        ]) == 0
        assert out.exists()
        report = out.with_suffix(".report.txt").read_text(encoding="utf-8")
        assert "报价行总数: 0" in report
