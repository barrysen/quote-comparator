"""阶段六 export 单测：明细 Sheet、标黄、失败原文 Sheet、空输入（M1 空报告）。"""

import json
from pathlib import Path

from openpyxl import load_workbook

from src.pipeline import export


def _quote(flags=None, confidence="high"):
    return {
        "source_file": "a.xlsx", "kind": "excel", "supplier_name": "供应商A",
        "lines": [{
            "item_name": "内六角螺钉", "spec": "M8×20", "quantity_raw": "5000",
            "quantity": 5000.0, "unit": "件", "unit_price_raw": "0.05", "unit_price": 0.05,
            "amount_raw": "250", "currency": "CNY", "tax_included": None,
            "lead_time_raw": "现货", "lead_time_days": None, "payment_terms": "",
            "remark": "", "confidence": confidence, "source_snippet": "...",
            "flags": flags or [],
        }],
        "warnings": [],
    }


def test_detail_sheet_and_yellow(tmp_path):
    out = tmp_path / "r.xlsx"
    good = _quote()
    bad = _quote(flags=["V2"], confidence="low")
    bad["source_file"] = "b.txt"
    bad["supplier_name"] = "供应商B"
    result = export.run({"quotes": [good, bad], "findings": [], "output": str(out)}, {})

    wb = load_workbook(out)
    assert "明细数据" in wb.sheetnames
    assert "统计与异常" in wb.sheetnames
    ws = wb["明细数据"]
    assert ws.max_row == 3  # 表头 + 2 行
    no_fill = ws.cell(row=2, column=1).fill.start_color.rgb
    yes_fill = ws.cell(row=3, column=1).fill.start_color.rgb
    assert str(yes_fill).endswith("FFF3CD")
    assert not str(no_fill).endswith("FFF3CD")
    # result.json / report.txt 同输出
    assert Path(result["result_json"]).exists()
    assert Path(result["report"]).exists()
    data = json.loads(Path(result["result_json"]).read_text(encoding="utf-8"))
    assert len(data["quotes"]) == 2


def test_empty_input_generates_empty_report(tmp_path):
    """M1 完成标志：空文件夹跑通、生成空报告。"""
    out = tmp_path / "empty.xlsx"
    result = export.run({"quotes": [], "findings": [], "output": str(out)}, {})
    wb = load_workbook(out)
    assert wb["明细数据"].max_row == 1  # 仅表头
    report = Path(result["report"]).read_text(encoding="utf-8")
    assert "报价行总数: 0" in report


def test_failed_raw_sheet(tmp_path):
    out = tmp_path / "r.xlsx"
    export.run({
        "quotes": [], "findings": [],
        "failed_files": [{"file": "bad.xlsx", "reason": "LLM 解析失败"}],
        "failed_raw": [{"file": "bad.xlsx", "raw_text": "原始内容"}],
        "output": str(out),
    }, {})
    wb = load_workbook(out)
    assert "解析失败原文" in wb.sheetnames
    ws = wb["解析失败原文"]
    assert ws.cell(row=2, column=2).value == "原始内容"
    assert str(ws.cell(row=2, column=2).fill.start_color.rgb).endswith("FFF3CD")


# ---------- M3：比价总表标色 ----------

def _cline(name, spec, price, lead=None, flags=None):
    return {
        "item_name": name, "spec": spec, "quantity_raw": "", "quantity": None,
        "unit": "件", "unit_price_raw": str(price), "unit_price": float(price),
        "amount_raw": "", "currency": "CNY", "tax_included": None,
        "lead_time_raw": "", "lead_time_days": lead, "payment_terms": "",
        "remark": "", "confidence": "high", "source_snippet": "x", "flags": flags or [],
    }


def _cquote(supplier, lines):
    return {"source_file": f"{supplier}.xlsx", "kind": "excel",
            "supplier_name": supplier, "lines": lines, "warnings": []}


def _fill_of(ws, row, col):
    rgb = ws.cell(row=row, column=col).fill.start_color.rgb
    return str(rgb)[-6:] if rgb else ""


def test_compare_sheet_colors(tmp_path):
    """预设比价情形全部标色正确：绿=最低价 红=超均价15% 灰=缺报 黄=低置信(最高优先) 橙=交期异常。"""
    from src.pipeline import compare

    quotes = [
        _cquote("A", [_cline("X", "s", 100, lead=5), _cline("Y", "s", 50, lead=3),
                      _cline("Z", "s", 10, flags=["V2"])]),
        _cquote("B", [_cline("X", "s", 120, lead=7), _cline("Y", "s", 52, lead=4),
                      _cline("Z", "s", 12)]),
        _cquote("C", [_cline("X", "s", 140, lead=20)]),
    ]
    config = {"compare": {"price_deviation_threshold": 0.15, "lead_time_factor": 2.0},
              "prompts": {"match_items": "config/prompts/match_items.md"}}
    cmp = compare.run({"quotes": quotes}, config, llm=None)
    assert len(cmp.groups) == 3

    out = tmp_path / "c.xlsx"
    export.run({"quotes": quotes, "findings": [], "comparison": cmp.model_dump(mode="json"),
                "output": str(out)}, config)
    wb = load_workbook(out)
    assert wb.sheetnames[0] == "比价总表"
    ws = wb["比价总表"]

    rows = {ws.cell(row=r, column=1).value: r for r in range(3, ws.max_row + 1)}
    rx, ry, rz = rows["X"], rows["Y"], rows["Z"]
    # 列布局：A(1) B(2) | A家 3-6 | B家 7-10 | C家 11-14 | 辅助 15-18
    # 绿：X 组最低价 A=100
    assert _fill_of(ws, rx, 3) == "C6EFCE"
    # 红：X 组 C=140，偏离均价 120 约 16.7%
    assert _fill_of(ws, rx, 11) == "FFC7CE"
    # 灰：Y 组 C 缺报
    assert ws.cell(row=ry, column=11).value == "缺报"
    assert _fill_of(ws, ry, 11) == "D9D9D9"
    # 黄优先：Z 组 A 虽是最低价，但有 V2 标记 → 黄覆盖绿
    assert _fill_of(ws, rz, 3) == "FFF3CD"
    # 橙：X 组 C 交期 20 天 > 中位数 7 × 2
    assert _fill_of(ws, rx, 13) == "FFE0B2"
    # 均价辅助列
    assert ws.cell(row=rx, column=15).value == 120.0

    # 缺报清单写入统计 Sheet
    ws2 = wb["统计与异常"]
    texts = [ws2.cell(row=r, column=1).value for r in range(1, ws2.max_row + 1)]
    assert "缺报清单" in texts

