"""阶段二：原始内容提取 —— 任何格式统一转成"保留结构信息的文本"。

不做业务理解。M2 实现 excel / text / pdf_digital；pdf_scanned / image 的
OCR 管线在 M4 实现，当前产出警告并跳过（不中断批处理）。
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from src.models import ClassifyResult, RawContent


def run(input: ClassifyResult | dict, config: dict) -> RawContent:
    c = ClassifyResult.model_validate(input)
    path = Path(c.file)
    name = path.name
    if c.kind == "excel":
        return RawContent(source_file=name, kind=c.kind, raw_text=_excel_to_text(path))
    if c.kind == "text":
        text, warnings = _read_text(path)
        return RawContent(
            source_file=name, kind=c.kind, raw_text=text, extraction_warnings=warnings
        )
    if c.kind == "pdf_digital":
        text, warnings = _pdf_digital_to_text(path)
        return RawContent(
            source_file=name, kind=c.kind, raw_text=text, extraction_warnings=warnings
        )
    if c.kind in ("pdf_scanned", "image"):
        return _ocr_extract(path, c, config)
    return RawContent(
        source_file=name,
        kind=c.kind,
        raw_text="",
        extraction_warnings=[f"不支持的格式，未提取: {c.detail}"],
    )


def _ocr_extract(path: Path, c: ClassifyResult, config: dict) -> RawContent:
    """pdf_scanned / image：转图 → OpenCV 矫正 → OCR，低置信行记入 low_ocr_lines。"""
    from src.pipeline import ocr

    name = path.name
    try:
        if c.kind == "image":
            lines, low = ocr.run_ocr_on_image(path, config)
        else:
            lines, low = ocr.run_ocr_on_pdf(path, config)
    except ocr.OCRUnavailable as e:
        return RawContent(
            source_file=name, kind=c.kind, raw_text="",
            extraction_warnings=[f"OCR 引擎不可用: {e}（安装 paddleocr 或 rapidocr-onnxruntime 后重跑）"],
        )
    except Exception as e:  # noqa: BLE001 —— 单文件失败不中断批处理
        return RawContent(
            source_file=name, kind=c.kind, raw_text="",
            extraction_warnings=[f"OCR 处理失败: {e}"],
        )
    warnings = [] if lines else ["OCR 未识别到任何文字"]
    if low:
        warnings.append(f"{len(low)} 行 OCR 置信度低于阈值，相关报价行将标黄")
    return RawContent(
        source_file=name, kind=c.kind, raw_text="\n".join(lines),
        extraction_warnings=warnings, low_ocr_lines=low,
    )


# ---------- excel ----------

def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "是" if v else "否"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).replace("|", "｜").replace("\n", " ").strip()


def _rows_to_markdown(rows: list[list[str]]) -> str:
    """单元格原值转 Markdown 表格；首行作表头，空单元格留空。"""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    lines = ["| " + " | ".join(rows[0]) + " |"]
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for r in rows[1:]:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def _excel_to_text(path: Path) -> str:
    if path.suffix.lower() == ".xls":
        return _xls_to_text(path)
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    parts = []
    for ws in wb.worksheets:
        rows = [[_cell_str(v) for v in row] for row in ws.iter_rows(values_only=True)]
        while rows and all(v == "" for v in rows[-1]):  # 去尾部全空行
            rows.pop()
        if not rows:
            continue
        parts.append(f"Sheet: {ws.title}\n" + _rows_to_markdown(rows))
    return "\n\n".join(parts)


def _xls_to_text(path: Path) -> str:
    import xlrd

    book = xlrd.open_workbook(str(path))
    parts = []
    for sheet in book.sheets():
        rows = [
            [_cell_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        while rows and all(v == "" for v in rows[-1]):
            rows.pop()
        if rows:
            parts.append(f"Sheet: {sheet.name}\n" + _rows_to_markdown(rows))
    return "\n\n".join(parts)


# ---------- text ----------

def _read_text(path: Path) -> tuple[str, list[str]]:
    import chardet

    raw = path.read_bytes()
    if not raw:
        return "", ["空文件"]
    detected = chardet.detect(raw)
    enc = detected.get("encoding") or "utf-8"
    try:
        return raw.decode(enc), []
    except (UnicodeDecodeError, LookupError):
        return raw.decode("utf-8", errors="replace"), [
            f"检测编码 {enc} 解码失败，已按 utf-8 容错（可能有乱码）"
        ]


# ---------- pdf_digital ----------

def _pdf_digital_to_text(path: Path) -> tuple[str, list[str]]:
    import pdfplumber

    warnings: list[str] = []
    parts: list[str] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                if text:
                    parts.append(f"[第{i}页 文本]\n{text}")
                for table in page.extract_tables() or []:
                    rows = [[_cell_str(v) for v in row] for row in table if row]
                    if rows:
                        parts.append(f"[第{i}页 表格]\n" + _rows_to_markdown(rows))
    except Exception as e:  # noqa: BLE001
        warnings.append(f"PDF 提取失败: {e}")
    if not parts and not warnings:
        warnings.append("PDF 未提取到文本（可能实为扫描件）")
    return "\n\n".join(parts), warnings
