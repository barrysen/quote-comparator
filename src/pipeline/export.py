"""阶段六：Excel 导出。

Sheet 结构（设计方案 §3.6）：
- 比价总表：行=物料，列=供应商（每组 单价/数量/交期/置信度），含均价/最低价/最高价辅助列（M3）
- 明细数据：长表，一行一条报价记录，含全部提取字段 + 来源文件 + 原文片段（溯源）
- 统计与异常：文件处理统计、失败清单、校验发现、缺报清单
- 解析失败原文：LLM 持续失败的文件，原文原样导出、整体标黄

比价标色规则（§3.5）：
| 情形 | 颜色 |
| 该物料最低价（≥2 家报价时） | 绿 C6EFCE |
| 高于均价 15% 以上 | 红 FFC7CE |
| 缺报 | 灰 D9D9D9 |
| 任何低置信度/校验未过/待确认匹配 | 黄 FFF3CD（优先级最高，覆盖其他颜色） |
| 交期为该物料最长且超过中位数 2 倍 | 橙 FFE0B2（交期单元格） |
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models import Comparison, Finding, SupplierQuote

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
GRAY = PatternFill("solid", fgColor="D9D9D9")
YELLOW = PatternFill("solid", fgColor="FFF3CD")
ORANGE = PatternFill("solid", fgColor="FFE0B2")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)

DETAIL_COLUMNS = [
    "来源文件", "供应商", "物料名称", "规格型号", "数量(原文)", "数量", "单位",
    "单价(原文)", "单价", "金额(原文)", "币种", "含税", "交期(原文)", "交期(天)",
    "付款方式", "备注", "置信度", "标记", "原文片段",
]
SUPPLIER_SUBCOLS = ["单价", "数量", "交期", "置信度"]


def run(input: dict, config: dict) -> dict:
    quotes = [SupplierQuote.model_validate(q) for q in input.get("quotes", [])]
    findings = [Finding.model_validate(f) for f in input.get("findings", [])]
    failed_files = input.get("failed_files", [])
    failed_raw = input.get("failed_raw", [])
    comparison = (
        Comparison.model_validate(input["comparison"]) if input.get("comparison") else None
    )
    out_path = Path(input["output"])
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    first = wb.active
    if comparison and comparison.groups:
        first.title = "比价总表"
        _write_compare(first, comparison, config)
        _write_detail(wb.create_sheet("明细数据"), quotes)
    else:
        first.title = "明细数据"
        _write_detail(first, quotes)
    _write_stats(wb.create_sheet("统计与异常"), quotes, findings, failed_files, comparison)
    if failed_raw:
        _write_failed_raw(wb.create_sheet("解析失败原文"), failed_raw)
    wb.save(out_path)

    json_path = out_path.with_suffix(".result.json")
    json_path.write_text(
        json.dumps(
            {
                "quotes": [q.model_dump(mode="json") for q in quotes],
                "findings": [f.model_dump(mode="json") for f in findings],
                "comparison": comparison.model_dump(mode="json") if comparison else None,
                "failed_files": failed_files,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    report_path = out_path.with_suffix(".report.txt")
    report_path.write_text(
        _build_report(quotes, findings, failed_files, comparison), encoding="utf-8"
    )
    return {"xlsx": str(out_path), "result_json": str(json_path), "report": str(report_path)}


# ---------- 比价总表（M3） ----------

def _write_compare(ws, comparison: Comparison, config: dict) -> None:
    cmp_cfg = config.get("compare", {})
    threshold = float(cmp_cfg.get("price_deviation_threshold", 0.15))
    lead_factor = float(cmp_cfg.get("lead_time_factor", 2.0))

    suppliers = comparison.suppliers
    # 表头：两行
    ws.merge_cells("A1:A2")
    ws["A1"] = "物料名称"
    ws.merge_cells("B1:B2")
    ws["B1"] = "规格型号"
    for si, supplier in enumerate(suppliers):
        c0 = 3 + si * len(SUPPLIER_SUBCOLS)
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + len(SUPPLIER_SUBCOLS) - 1)
        ws.cell(row=1, column=c0, value=supplier)
        for j, sub in enumerate(SUPPLIER_SUBCOLS):
            ws.cell(row=2, column=c0 + j, value=sub)
    aux_start = 3 + len(suppliers) * len(SUPPLIER_SUBCOLS)
    for j, title in enumerate(["均价", "最低价", "最高价", "备注"]):
        col = aux_start + j
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        ws.cell(row=1, column=col, value=title)
    for row in ws.iter_rows(min_row=1, max_row=2, max_col=aux_start + 3):
        for cell in row:
            cell.font = HEADER_FONT

    for group in comparison.groups:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=group.item_name)
        ws.cell(row=row, column=2, value=group.spec)
        priced = [e for e in group.entries if e.line.unit_price is not None]
        multi = len(group.entries) >= 2

        for si, supplier in enumerate(suppliers):
            c0 = 3 + si * len(SUPPLIER_SUBCOLS)
            entry = next((e for e in group.entries if e.supplier_name == supplier), None)
            cells = [ws.cell(row=row, column=c0 + j) for j in range(len(SUPPLIER_SUBCOLS))]
            if entry is None:
                if multi:  # ≥2 家报价而此家未报 → 缺报标灰
                    cells[0].value = "缺报"
                    for cell in cells:
                        cell.fill = GRAY
                continue
            line = entry.line
            cells[0].value = line.unit_price if line.unit_price is not None else (line.unit_price_raw or "")
            cells[1].value = line.quantity if line.quantity is not None else (line.quantity_raw or "")
            cells[2].value = (
                line.lead_time_days if line.lead_time_days is not None else (line.lead_time_raw or "")
            )
            cells[3].value = line.confidence + (f" ({','.join(line.flags)})" if line.flags else "")

            yellow = line.confidence == "low" or bool(line.flags) or group.review
            if yellow:  # 黄色优先级最高，覆盖其他颜色
                for cell in cells:
                    cell.fill = YELLOW
                continue
            if len(priced) >= 2 and line.unit_price is not None:
                if line.unit_price == group.min_price:
                    cells[0].fill = GREEN
                elif group.deviations.get(supplier, 0) > threshold:
                    cells[0].fill = RED
            leads = [e.line.lead_time_days for e in group.entries if e.line.lead_time_days is not None]
            if (
                len(leads) >= 2
                and line.lead_time_days is not None
                and group.max_lead_days is not None
                and group.median_lead_days is not None
                and line.lead_time_days == group.max_lead_days
                and group.max_lead_days > lead_factor * group.median_lead_days
            ):
                cells[2].fill = ORANGE

        for j, val in enumerate([group.avg_price, group.min_price, group.max_price]):
            if val is not None:
                ws.cell(row=row, column=aux_start + j, value=round(val, 4))
        if group.notes:
            ws.cell(row=row, column=aux_start + 3, value="；".join(group.notes))

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(aux_start + 3)}{ws.max_row}"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    for si in range(len(suppliers)):
        for j, width in enumerate([10, 8, 8, 12]):
            ws.column_dimensions[get_column_letter(3 + si * 4 + j)].width = width
    for j, width in enumerate([10, 10, 10, 40]):
        ws.column_dimensions[get_column_letter(aux_start + j)].width = width


# ---------- 明细数据 ----------

def _is_yellow(line) -> bool:
    return line.confidence == "low" or bool(line.flags)


def _write_detail(ws, quotes: list[SupplierQuote]) -> None:
    ws.append(DETAIL_COLUMNS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for quote in quotes:
        for line in quote.lines:
            ws.append([
                quote.source_file,
                quote.supplier_name,
                line.item_name,
                line.spec,
                line.quantity_raw,
                line.quantity,
                line.unit,
                line.unit_price_raw,
                line.unit_price,
                line.amount_raw,
                line.currency,
                "" if line.tax_included is None else ("含税" if line.tax_included else "未税"),
                line.lead_time_raw,
                line.lead_time_days,
                line.payment_terms,
                line.remark,
                line.confidence,
                ",".join(line.flags),
                line.source_snippet,
            ])
            if _is_yellow(line):
                for cell in ws[ws.max_row]:
                    cell.fill = YELLOW
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLUMNS))}1"
    for i, width in enumerate(
        [16, 18, 16, 14, 10, 8, 6, 12, 8, 10, 6, 6, 10, 8, 12, 12, 8, 8, 40], start=1
    ):
        ws.column_dimensions[get_column_letter(i)].width = width


# ---------- 统计与异常 ----------

def _write_stats(
    ws,
    quotes: list[SupplierQuote],
    findings: list[Finding],
    failed_files: list[dict],
    comparison: Comparison | None,
) -> None:
    row = 1

    def write_row(values, bold=False):
        nonlocal row
        ws.append(values)
        if bold:
            for cell in ws[row]:
                cell.font = HEADER_FONT
        row += 1

    ws.cell(row=row, column=1, value="文件处理统计").font = TITLE_FONT
    row += 1
    write_row(["文件", "类型", "供应商", "报价行数", "警告数", "状态"], bold=True)
    for q in quotes:
        status = "失败" if any(w.startswith("LLM_FAILED") for w in q.warnings) else (
            "警告" if q.warnings else "正常"
        )
        write_row([q.source_file, q.kind, q.supplier_name, len(q.lines), len(q.warnings), status])
    row += 1

    ws.cell(row=row, column=1, value="失败清单").font = TITLE_FONT
    row += 1
    write_row(["文件", "原因"], bold=True)
    for f in failed_files:
        write_row([f.get("file", ""), f.get("reason", "")])
    row += 1

    if comparison is not None:
        ws.cell(row=row, column=1, value="缺报清单").font = TITLE_FONT
        row += 1
        write_row(["物料", "缺报供应商"], bold=True)
        for m in comparison.missing:
            write_row([m.get("item", ""), m.get("supplier", "")])
        row += 1

    v3_findings = [f for f in findings if f.rule == "V3"]
    ws.cell(row=row, column=1, value="金额不符清单（V3）").font = TITLE_FONT
    row += 1
    write_row(["对象定位", "说明"], bold=True)
    for f in v3_findings:
        write_row([f.target, f.message])
    row += 1

    if comparison is not None and comparison.notes:
        ws.cell(row=row, column=1, value="比价说明").font = TITLE_FONT
        row += 1
        for note in comparison.notes:
            write_row([note])
        row += 1

    ws.cell(row=row, column=1, value="校验发现（V1~V7 / H1）").font = TITLE_FONT
    row += 1
    write_row(["对象定位", "规则号", "说明"], bold=True)
    for f in findings:
        write_row([f.target, f.rule, f.message])
    for col, width in zip("ABC", [28, 20, 60]):
        ws.column_dimensions[col].width = width


# ---------- 解析失败原文 ----------

def _write_failed_raw(ws, failed_raw: list[dict]) -> None:
    """LLM 持续失败的文件：原始文本原样导出到单独 Sheet，整体标黄。"""
    ws.append(["文件", "原始文本（未解析，请人工处理）"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = YELLOW
    for item in failed_raw:
        ws.append([item.get("file", ""), item.get("raw_text", "")])
        for cell in ws[ws.max_row]:
            cell.fill = YELLOW
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100


# ---------- report.txt ----------

def _build_report(
    quotes: list[SupplierQuote],
    findings: list[Finding],
    failed_files: list[dict],
    comparison: Comparison | None,
) -> str:
    lines = [
        "报价单提取比价工具 · 处理报告",
        "=" * 40,
        f"处理文件数: {len(quotes) + len([f for f in failed_files if f.get('reason', '').startswith('不支持')])}",
        f"成功解析: {sum(1 for q in quotes if q.lines)}",
        f"失败: {len(failed_files)}",
        f"报价行总数: {sum(len(q.lines) for q in quotes)}",
    ]
    if comparison is not None:
        review = sum(1 for g in comparison.groups if g.review)
        lines += [
            "",
            "比价分析:",
            f"  供应商: {len(comparison.suppliers)} 家（{'、'.join(comparison.suppliers)}）",
            f"  物料组: {len(comparison.groups)} 组",
            f"  缺报: {len(comparison.missing)} 条",
            f"  待人工确认匹配: {review} 组",
        ]
        for note in comparison.notes:
            lines.append(f"    - {note}")
    lines.append("")
    for q in quotes:
        lines.append(
            f"[{q.source_file}] 供应商: {q.supplier_name}，{len(q.lines)} 行，{len(q.warnings)} 条警告"
        )
        for w in q.warnings:
            lines.append(f"    - {w}")
    if failed_files:
        lines.append("")
        lines.append("失败清单:")
        for f in failed_files:
            lines.append(f"    - {f.get('file')}: {f.get('reason')}")
    lines.append("")
    lines.append(f"校验发现: {len(findings)} 条")
    for f in findings:
        lines.append(f"    [{f.rule}] {f.target}: {f.message}")
    lines.append("")
    lines.append("提示: 标黄（低置信/校验未过/待确认匹配）的单元格需要人工复核后再使用。")
    return "\n".join(lines)
